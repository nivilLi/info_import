from openpyxl import *
import pyautogui
from collections import Iterable
# # 创建一个 workbook
# wb = Workbook()
# 获取被激活的 worksheet
# ws = wb.active
# # 设置单元格内容
# ws['A1'] = 42
# # 设置一行内容
# ws.append([1, 2, 3])
# # python 数据类型可以被自动转换
# import datetime
# ws['A2'] = datetime.datetime.now()
# # 保存 Excel 文件
# wb.save("d:\\sample.xlsx")
wb = load_workbook('D:\\wenjian\\WeChat Files\\wxid_72kpxoqox11g22\\FileStorage\\File\\2022-12\\李志超-黄石村乡村建设信息.xlsx')
ws = wb.active
rows = ws.iter_rows()
i = 0
# print(rows)
for row in rows:
    i = i + 1
    if i < 4:
        continue
    for cell in row:
        print(cell.value)